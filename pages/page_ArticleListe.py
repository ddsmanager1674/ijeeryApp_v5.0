import customtkinter
from format_utils import format_entier
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import psycopg2
import json
import os
import subprocess
import sys
from PIL import Image, ImageTk
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime
from resource_utils import get_config_path, safe_file_read

# ── Thème iJeery ──────────────────────────────────────────────────────────────
try:
    from app_theme import Colors, Fonts, styled, Theme
    _T = True
except ImportError:
    _T = False


class _C:
    """Fallback couleurs si app_theme absent."""
    MIDNIGHT       = "#2C3E50"
    BG_PAGE        = "#ECF0F1"
    BG_CARD        = "#FFFFFF"
    BG_HEADER      = "#2C3E50"
    BG_INPUT       = "#F4F6F8"
    PRIMARY        = "#3498DB"
    PRIMARY_HOVER  = "#2980B9"
    SUCCESS        = "#2ECC71"
    SUCCESS_DARK   = "#27AE60"
    SUCCESS_LIGHT  = "#D5F5E3"
    SUCCESS_TEXT   = "#1E8449"
    DANGER         = "#E74C3C"
    DANGER_DARK    = "#C0392B"
    DANGER_LIGHT   = "#FADBD8"
    DANGER_TEXT    = "#922B21"
    WARNING        = "#F39C12"
    WARNING_LIGHT  = "#FEF9E7"
    WARNING_TEXT   = "#9A6A00"
    INFO           = "#1ABC9C"
    INFO_DARK      = "#16A085"
    PREMIUM        = "#9B59B6"
    PREMIUM_DARK   = "#8E44AD"
    TEXT_PRIMARY   = "#2C3E50"
    TEXT_SECONDARY = "#5D6D7E"
    TEXT_MUTED     = "#95A5A6"
    BORDER         = "#D5D8DC"
    DIVIDER        = "#E8EAED"
    SILVER         = "#BDC3C7"
    CLOUDS         = "#ECF0F1"


C = Colors if _T else _C


# ====================================================================
# GESTION DES IMPORTATIONS (Compatible VSCode ET app_main)
# ====================================================================

def import_page_info_article():
    """Tente d'importer PageInfoArticle avec différentes stratégies"""
    try:
        from pages.page_infoArticle import PageInfoArticle
        print("✓ PageInfoArticle importée depuis pages.page_infoArticle")
        return PageInfoArticle
    except ImportError:
        try:
            from page_infoArticle import PageInfoArticle
            print("✓ PageInfoArticle importée depuis page_infoArticle")
            return PageInfoArticle
        except ImportError as e:
            print(f"❌ Erreur d'import PageInfoArticle: {e}")

            class PageInfoArticleFallback(customtkinter.CTkFrame):
                def __init__(self, master, db_conn=None, session_data=None, initial_idarticle=None):
                    super().__init__(master)
                    self.pack(fill="both", expand=True)
                    error_frame = customtkinter.CTkFrame(self, fg_color=C.DANGER_LIGHT)
                    error_frame.pack(fill="both", expand=True, padx=20, pady=20)
                    customtkinter.CTkLabel(
                        error_frame,
                        text="❌ Erreur de chargement",
                        font=customtkinter.CTkFont(
                            family="Roboto" if _T else "Segoe UI", size=20, weight="bold"),
                        text_color=C.DANGER_TEXT
                    ).pack(pady=20)
                    customtkinter.CTkLabel(
                        error_frame,
                        text=f"Impossible de charger PageInfoArticle\nArticle ID: {initial_idarticle}",
                        font=customtkinter.CTkFont(
                            family="Roboto" if _T else "Segoe UI", size=12),
                        text_color=C.TEXT_SECONDARY
                    ).pack(pady=10)

            return PageInfoArticleFallback


PageInfoArticle = import_page_info_article()


# ====================================================================
# 1. Style Treeview — thème iJeery
# ====================================================================

def configure_treeview_style(root):
    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except Exception:
        pass

    style.configure("Liste.Treeview",
                    background=C.BG_CARD,
                    foreground=C.TEXT_PRIMARY,
                    fieldbackground=C.BG_CARD,
                    borderwidth=0,
                    font=("Roboto" if _T else "Segoe UI", 10),
                    rowheight=24)

    style.configure("Liste.Treeview.Heading",
                    background=C.BG_HEADER,
                    foreground="#FFFFFF",
                    font=("Roboto" if _T else "Segoe UI", 10, "bold"),
                    relief="flat",
                    padding=(6, 4))

    style.map("Liste.Treeview",
              background=[("selected", C.PRIMARY)],
              foreground=[("selected", "#FFFFFF")])

    root.option_add('*Treeview*highlightThickness', 0)


# ====================================================================
# 2. Classe principale
# ====================================================================

class page_listeArticle(customtkinter.CTkFrame):

    def __init__(self, master, db_conn=None, session_data=None, **kwargs):
        super().__init__(master, fg_color=C.BG_PAGE, **kwargs)

        self.db_conn      = db_conn
        self.session_data = session_data

        try:
            configure_treeview_style(master)
        except Exception as e:
            print(f"Erreur style Treeview: {e}")

        self.grid_rowconfigure(2, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.all_data = []
        self.show_photo_column = tk.BooleanVar(value=False)
        self._photo_image_refs = {}
        self._photo_thumb_cache = {}

        self._build_header()
        self._build_filters()
        self._build_treeview()
        self._build_footer()

        self.load_data()

    # ── helpers ──────────────────────────────────────────────────────────────
    def _f(self, size=11, weight="normal"):
        return customtkinter.CTkFont(
            family="Roboto" if _T else "Segoe UI",
            size=size, weight=weight)

    # ── En-tête ───────────────────────────────────────────────────────────────
    def _build_header(self):
        hdr = customtkinter.CTkFrame(self, fg_color=C.BG_HEADER, corner_radius=0)
        hdr.grid(row=0, column=0, sticky="ew")

        customtkinter.CTkLabel(
            hdr, text="Liste des Articles",
            font=self._f(18, "bold"),
            text_color="#FFFFFF"
        ).pack(side="left", padx=16, pady=10)

        # Compteur côté droit dans le header
        self._lbl_count_hdr = customtkinter.CTkLabel(
            hdr, text="",
            font=self._f(9), text_color="#AAAAAA")
        self._lbl_count_hdr.pack(side="right", padx=16)

    # ── Barre filtres + boutons actions ──────────────────────────────────────
    def _build_filters(self):
        panel = customtkinter.CTkFrame(self, fg_color=C.BG_CARD, corner_radius=8)
        panel.grid(row=1, column=0, sticky="ew", padx=12, pady=6)

        inner = customtkinter.CTkFrame(panel, fg_color="transparent")
        inner.pack(fill="x", padx=10, pady=8)

        # ── Recherche ────────────────────────────────────────────────────────
        customtkinter.CTkLabel(
            inner, text="🔍", font=self._f(13),
            text_color=C.TEXT_MUTED
        ).pack(side="left", padx=(0, 4))

        self.entry_search = customtkinter.CTkEntry(
            inner,
            placeholder_text="Rechercher par Code, Désignation, Unité ou Catégorie…",
            width=320, height=30,
            fg_color=C.BG_INPUT, border_color=C.BORDER,
            text_color=C.TEXT_PRIMARY,
            font=self._f(10))
        self.entry_search.pack(side="left", padx=(0, 6))
        self.entry_search.bind('<KeyRelease>', lambda e: self.filter_data())

        customtkinter.CTkButton(
            inner, text="Réinitialiser",
            command=self.reset_filters,
            fg_color="transparent", hover_color=C.DIVIDER,
            text_color=C.TEXT_SECONDARY,
            border_width=1, border_color=C.BORDER,
            height=30, width=100,
            font=self._f(10)
        ).pack(side="left", padx=(0, 16))

        # Séparateur vertical
        customtkinter.CTkFrame(
            inner, width=1, height=22, fg_color=C.BORDER
        ).pack(side="left", padx=(0, 12))

        # ── Boutons actions ───────────────────────────────────────────────────
        customtkinter.CTkButton(
            inner, text="📁  Gérer Articles",
            command=self.open_new_article,
            fg_color=C.PRIMARY, hover_color=C.PRIMARY_HOVER,
            text_color="#FFFFFF",
            height=30, width=150,
            font=self._f(10, "bold")
        ).pack(side="left", padx=(0, 6))

        customtkinter.CTkButton(
            inner, text="🗂  Gérer Catégories",
            command=self.open_new_category,
            fg_color=C.TEXT_SECONDARY, hover_color=C.MIDNIGHT,
            text_color="#FFFFFF",
            height=30, width=160,
            font=self._f(10, "bold")
        ).pack(side="left", padx=(0, 6))

        customtkinter.CTkButton(
            inner, text="📊  Exporter Excel",
            command=self.export_to_excel,
            fg_color=C.SUCCESS_DARK, hover_color=C.SUCCESS,
            text_color="#FFFFFF",
            height=30, width=150,
            font=self._f(10, "bold")
        ).pack(side="right", padx=(6, 0))

    # ── Treeview ──────────────────────────────────────────────────────────────
    def _build_treeview(self):
        tbl = customtkinter.CTkFrame(self, fg_color=C.BG_CARD, corner_radius=8)
        tbl.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 4))
        tbl.grid_rowconfigure(0, weight=1)
        tbl.grid_columnconfigure(0, weight=1)

        columns = ("ID", "Code", "Designation", "Unite",
                   "Quantite", "Poids", "Categorie")

        self.tree = ttk.Treeview(
            tbl, columns=columns, show="headings",
            style="Liste.Treeview", height=18)

        self.tree.tag_configure("even", background=C.BG_CARD)
        self.tree.tag_configure("odd",  background="#F0F4F8")

        # En-têtes
        headers = {
            "ID":          ("ID",                   0,   False),
            "Code":        ("Code Article",        120,  True),
            "Designation": ("Désignation",         300,  True),
            "Unite":       ("Unité",                90,  True),
            "Quantite":    ("Quantité",            100,  True),
            "Poids":       ("Poids",               100,  True),
            "Categorie":   ("Catégorie",           180,  True),
        }
        for col, (label, w, stretch) in headers.items():
            self.tree.heading(col, text=label)
            anchor = "center" if col in ("Quantite", "Poids", "Unite") else "w"
            self.tree.column(col, width=w, stretch=stretch,
                             anchor=anchor, minwidth=0 if w == 0 else 40)
            if w == 0:
                self.tree.column(col, width=0, stretch=False, minwidth=0)
        self.tree.heading("#0", text="Photo")
        self.tree.column("#0", width=0, stretch=False, minwidth=0, anchor="center")

        sy = customtkinter.CTkScrollbar(tbl, orientation="vertical",
                                        command=self.tree.yview)
        sx = customtkinter.CTkScrollbar(tbl, orientation="horizontal",
                                        command=self.tree.xview)
        self.tree.configure(yscrollcommand=sy.set, xscrollcommand=sx.set)

        self.tree.grid(row=0, column=0, sticky="nsew", padx=(6, 0), pady=(6, 0))
        sy.grid(row=0, column=1, sticky="ns", pady=(6, 0))
        sx.grid(row=1, column=0, sticky="ew", padx=(6, 0))

        self.tree.bind('<Double-Button-1>', self.on_double_click)
        self.tree.bind('<ButtonRelease-1>', self.on_single_click)

        # Badge d'export ajoute automatiquement via export_utils.enable_treeview_export_badge()

    # ── Footer ────────────────────────────────────────────────────────────────
    def _build_footer(self):
        footer = customtkinter.CTkFrame(self, fg_color="transparent")
        footer.grid(row=3, column=0, sticky="ew", padx=12, pady=(2, 8))

        self.label_count = customtkinter.CTkLabel(
            footer,
            text="0 article(s)",
            font=self._f(10, "bold"),
            text_color=C.PRIMARY)
        self.label_count.pack(side="left")

        self.chk_show_photos = customtkinter.CTkCheckBox(
            footer,
            text="Afficher colonne photo",
            variable=self.show_photo_column,
            onvalue=True,
            offvalue=False,
            command=self.toggle_photo_column,
            font=self._f(10),
            text_color=C.TEXT_SECONDARY,
        )
        self.chk_show_photos.pack(side="right")

    # ====================================================================
    # LOGIQUE MÉTIER — inchangée
    # ====================================================================

    def connect_db(self):
        if self.db_conn:
            return self.db_conn
        conn = None
        try:
            with open(get_config_path('config.json')) as f:
                config = json.load(f)
                db_config = config['database']
            conn = psycopg2.connect(
                host=db_config['host'],
                user=db_config['user'],
                password=db_config['password'],
                database=db_config['database'],
                port=db_config['port']
            )
            return conn
        except FileNotFoundError:
            messagebox.showerror("Erreur de configuration",
                                 "Fichier 'config.json' non trouvé.")
        except KeyError:
            messagebox.showerror("Erreur de configuration",
                                 "Clés de base de données manquantes dans 'config.json'.")
        except psycopg2.Error as err:
            messagebox.showerror("Erreur de connexion",
                                 f"Erreur de connexion à PostgreSQL : {err}")
        except Exception as err:
            messagebox.showerror("Erreur Inattendue",
                                 f"Une erreur inattendue est survenue : {err}")
        return None

    def fetch_articles_from_db(self):
        conn = self.connect_db()
        if conn is None:
            return []
        SQL_QUERY = """
        SELECT
            T2."idarticle",
            T1."codearticle",
            T2."designation",
            T1."designationunite",
            T1."qtunite",
            T1."poids",
            T3."designationcat"
        FROM
            tb_unite AS T1
        INNER JOIN
            tb_article AS T2 ON T1.idarticle = T2.idarticle
        INNER JOIN
            tb_categoriearticle AS T3 ON T2.idca = T3.idca
        ORDER BY T2."designation" ASC, T1."codearticle" ASC;
        """
        data = []
        try:
            with conn.cursor() as cur:
                cur.execute(SQL_QUERY)
                data = cur.fetchall()
        except psycopg2.Error as err:
            messagebox.showerror("Erreur SQL",
                                 f"Erreur lors de l'exécution de la requête : {err}")
        finally:
            if conn and conn != self.db_conn:
                conn.close()
        return data

    def _get_photos_folder(self):
        try:
            with open(get_config_path('config.json')) as f:
                config = json.load(f)
            configured_path = (
                config.get("photos_path")
                or config.get("photo_path")
                or config.get("photos", {}).get("path")
            )
            if configured_path:
                return configured_path
            server_cfg = config.get("server", {})
            server_ip = (
                server_cfg.get("ip")
                or config.get("server_ip")
                or config.get("ip_server")
                or config.get("database", {}).get("host")
                or "localhost"
            )
        except Exception:
            server_ip = "localhost"

        if str(server_ip).lower() in ("localhost", "127.0.0.1"):
            return os.path.join(os.getcwd(), "photos")
        return rf"\\{server_ip}\photos"

    def _normalize_article_id(self, article_id):
        if article_id is None:
            return ""
        s = str(article_id).strip()
        if not s:
            return ""
        if s.endswith(".0"):
            s = s[:-2]
        return s

    def _find_photo_path(self, idarticle):
        article_id = self._normalize_article_id(idarticle)
        if not article_id:
            return None
        photo_dir = self._get_photos_folder()
        for ext in (".jpg", ".jpeg", ".png", ".gif", ".bmp"):
            candidate = os.path.join(photo_dir, f"{article_id}{ext}")
            if os.path.exists(candidate):
                return candidate
        return None

    def _get_photo_thumbnail(self, idarticle, size=(42, 42)):
        article_id = self._normalize_article_id(idarticle)
        if not article_id:
            return None
        cache_key = (article_id, size)
        if cache_key in self._photo_thumb_cache:
            return self._photo_thumb_cache[cache_key]
        photo_path = self._find_photo_path(article_id)
        if not photo_path:
            self._photo_thumb_cache[cache_key] = None
            return None
        try:
            with Image.open(photo_path) as pil_img:
                img = pil_img.convert("RGB")
            img.thumbnail(size, Image.Resampling.LANCZOS)
            thumb = ImageTk.PhotoImage(img)
        except Exception:
            thumb = None
        self._photo_thumb_cache[cache_key] = thumb
        return thumb

    def _to_tree_row(self, row):
        return (row[0], row[1], row[2], row[3], row[4], row[5], row[6])

    def toggle_photo_column(self):
        if self.show_photo_column.get():
            self.tree.configure(show=("tree", "headings"))
            self.tree.column("#0", width=56, stretch=False, minwidth=56, anchor="center")
            self.tree.column("#0", anchor="center")
            ttk.Style().configure("Liste.Treeview", rowheight=48)
        else:
            self.tree.configure(show="headings")
            self.tree.column("#0", width=0, stretch=False, minwidth=0)
            ttk.Style().configure("Liste.Treeview", rowheight=24)
            self._photo_image_refs = {}
        self.filter_data()

    def _insert_rows_with_alternating_colors(self, rows):
        self._photo_image_refs = {}
        for index, row in enumerate(rows):
            tag = "even" if index % 2 == 0 else "odd"
            # Format Quantite (4ème colonne)
            formatted_row = list(self._to_tree_row(row))
            if len(formatted_row) > 4:
                formatted_row[4] = format_entier(formatted_row[4])

            image = None
            if self.show_photo_column.get():
                image = self._get_photo_thumbnail(row[0], size=(42, 42))

            if image is not None:
                iid = self.tree.insert('', 'end', values=formatted_row, tags=(tag,), image=image)
                self._photo_image_refs[iid] = image
            else:
                self.tree.insert('', 'end', values=formatted_row, tags=(tag,))

    def on_single_click(self, event):
        selection = self.tree.selection()
        if selection:
            self.tree.selection_set(selection)

    def on_double_click(self, event):
        selection = self.tree.selection()
        if selection:
            item = self.tree.item(selection[0])
            values = item['values']
            if values and len(values) > 0:
                idarticle = values[0]
                self.open_info_article(idarticle)

    def open_info_article(self, idarticle):
        try:
            info_window = customtkinter.CTkToplevel(self)
            info_window.title(f"Détails Article - ID: {idarticle}")
            info_window.geometry("900x600")
            if _T:
                Theme.apply_toplevel(info_window)
            info_window.grid_columnconfigure(0, weight=1)
            info_window.grid_rowconfigure(0, weight=1)
            page_frame = PageInfoArticle(
                master=info_window,
                db_conn=self.db_conn,
                session_data=self.session_data,
                initial_idarticle=str(idarticle)
            )
            page_frame.grid(row=0, column=0, sticky="nsew")
            info_window.update_idletasks()
            x = (info_window.winfo_screenwidth()  // 2) - (900 // 2)
            y = (info_window.winfo_screenheight() // 2) - (600 // 2)
            info_window.geometry(f"900x600+{x}+{y}")
            info_window.focus()
            info_window.lift()
        except Exception as e:
            messagebox.showerror(
                "Erreur",
                f"Impossible d'ouvrir la page d'information:\n\n{str(e)}\n\n"
                "Vérifiez que page_infoArticle.py est accessible.")
            import traceback; traceback.print_exc()

    def open_new_article(self):
        try:
            from pages.page_article import PageArticle
            article_window = customtkinter.CTkToplevel(self)
            article_window.title("Gérer Articles")
            article_window.geometry("700x700")
            if _T:
                Theme.apply_toplevel(article_window)
            article_window.transient(self.master)
            article_window.grab_set()
            article_window.grid_columnconfigure(0, weight=1)
            article_window.grid_rowconfigure(0, weight=1)
            try:
                page_frame = PageArticle(master=article_window,
                                         db_conn=self.db_conn,
                                         session_data=self.session_data)
            except TypeError:
                try:
                    page_frame = PageArticle(article_window,
                                             db_conn=self.db_conn,
                                             session_data=self.session_data)
                except TypeError:
                    page_frame = PageArticle(article_window)
            page_frame.grid(row=0, column=0, sticky="nsew")
            article_window.update_idletasks()
            x = (article_window.winfo_screenwidth()  // 2) - (600 // 2)
            y = (article_window.winfo_screenheight() // 2) - (700 // 2)
            article_window.geometry(f"600x700+{x}+{y}")
            article_window.focus()
            article_window.lift()
            article_window.wait_window()
            self.load_data()
        except ImportError as e:
            messagebox.showerror("Erreur d'import",
                                 f"Impossible d'importer PageArticle.\n\nErreur: {e}")
        except Exception as e:
            messagebox.showerror("Erreur",
                                 f"Impossible d'ouvrir la page article : {str(e)}")

    def open_new_category(self):
        try:
            from pages.page_categorieArticle import PageCategorieArticle
            category_window = customtkinter.CTkToplevel(self)
            category_window.title("Gérer Catégorie")
            category_window.geometry("400x400")
            if _T:
                Theme.apply_toplevel(category_window)
            category_window.transient(self.master)
            category_window.grab_set()
            category_window.grid_columnconfigure(0, weight=1)
            category_window.grid_rowconfigure(0, weight=1)
            try:
                page_frame = PageCategorieArticle(master=category_window,
                                                   db_conn=self.db_conn,
                                                   session_data=self.session_data)
            except TypeError:
                try:
                    page_frame = PageCategorieArticle(category_window,
                                                       db_conn=self.db_conn,
                                                       session_data=self.session_data)
                except TypeError:
                    page_frame = PageCategorieArticle(category_window)
            page_frame.grid(row=0, column=0, sticky="nsew")
            category_window.update_idletasks()
            x = (category_window.winfo_screenwidth()  // 2) - (400 // 2)
            y = (category_window.winfo_screenheight() // 2) - (400 // 2)
            category_window.geometry(f"400x400+{x}+{y}")
            category_window.focus()
            category_window.lift()
            category_window.wait_window()
            self.load_data()
        except ImportError as e:
            messagebox.showerror("Erreur d'import",
                                 f"Impossible d'importer PageCategorieArticle.\n\nErreur: {e}")
        except Exception as e:
            messagebox.showerror("Erreur",
                                 f"Impossible d'ouvrir la page catégorie : {str(e)}")

    def load_data(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        self._photo_thumb_cache = {}
        self.all_data = self.fetch_articles_from_db()
        if self.all_data:
            self._insert_rows_with_alternating_colors(self.all_data)
            self.update_count(len(self.all_data))
        else:
            self.tree.insert('', 'end',
                             values=("", "", "Aucun article trouvé", "", "", "", ""),
                             tags=("even",))
            self.update_count(0)

    def filter_data(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
        search_term = self.entry_search.get().lower().strip()
        if not search_term:
            self._insert_rows_with_alternating_colors(self.all_data)
            self.update_count(len(self.all_data))
            return
        filtered_data = []
        for row in self.all_data:
            searchable_text = f"{row[1]} {row[2]} {row[3]} {row[6]}".lower()
            if search_term in searchable_text:
                filtered_data.append(row)
        if filtered_data:
            self._insert_rows_with_alternating_colors(filtered_data)
            self.update_count(len(filtered_data))
        else:
            self.tree.insert('', 'end',
                             values=("", "", "Aucun résultat trouvé", "", "", "", ""),
                             tags=("even",))
            self.update_count(0)

    def reset_filters(self):
        self.entry_search.delete(0, 'end')
        self.load_data()

    def update_count(self, count):
        self.label_count.configure(text=f"{count} article(s)")
        self._lbl_count_hdr.configure(
            text=f"Actualisé : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

    def export_to_excel(self):
        items = self.tree.get_children()
        if not items:
            messagebox.showwarning("Aucune donnée", "Aucune donnée à exporter.")
            return
        first_item = self.tree.item(items[0])
        if not first_item['values'] or first_item['values'][2] in [
                "Aucun article trouvé", "Aucun résultat trouvé"]:
            messagebox.showwarning("Aucune donnée",
                                   "Aucune donnée valide à exporter.")
            return
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=f"Articles_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        if not filename:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Articles"
            headers = ["Code Article", "Désignation Article",
                       "Unité", "Quantité", "Poids", "Catégorie"]
            ws.append(headers)
            header_fill = PatternFill(start_color="2C3E50",
                                      end_color="2C3E50", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center",
                                           vertical="center")
            valid_items_count = 0
            for item in items:
                values = self.tree.item(item)['values']
                if values and values[1]:
                    ws.append([values[1], values[2], values[3], values[4], values[5], values[6]])
                    valid_items_count += 1
            column_widths = [15, 40, 12, 12, 12, 20]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width
            ws.append([])
            total_row = ws.max_row
            ws[f'A{total_row}'] = f"Total: {valid_items_count} articles"
            ws[f'A{total_row}'].font = Font(bold=True, size=11)
            wb.save(filename)
            messagebox.showinfo(
                "Export réussi",
                f"Les données ont été exportées avec succès vers:\n{filename}")
        except PermissionError:
            messagebox.showerror(
                "Erreur d'accès",
                "Le fichier est peut-être ouvert dans Excel. "
                "Veuillez le fermer et réessayer.")
        except Exception as e:
            messagebox.showerror("Erreur d'export",
                                 f"Une erreur est survenue lors de l'export:\n{e}")


# ====================================================================
# 3. Test standalone
# ====================================================================

if __name__ == "__main__":
    customtkinter.set_appearance_mode("light")
    customtkinter.set_default_color_theme("blue")

    class App(customtkinter.CTk):
        def __init__(self):
            super().__init__()
            if _T:
                Theme.apply(self)
            self.title("iJeery — Liste des Articles")
            self.geometry("1200x700")
            self.grid_rowconfigure(0, weight=1)
            self.grid_columnconfigure(0, weight=1)
            page_listeArticle(
                master=self,
                db_conn=None,
                session_data=None
            ).grid(row=0, column=0, sticky="nsew")

    App().mainloop()

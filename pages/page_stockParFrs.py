import customtkinter as ctk
from tkinter import ttk, messagebox
import psycopg2
import json
from datetime import datetime
import threading
from resource_utils import get_config_path, safe_file_read

# ── Thème iJeery ──────────────────────────────────────────────────────────────
try:
    from app_theme import Colors, Fonts, styled, Theme
    _T = True
except ImportError:
    _T = False


class _C:
    MIDNIGHT       = "#2C3E50"
    BG_PAGE        = "#ECF0F1"
    BG_CARD        = "#FFFFFF"
    BG_HEADER      = "#2C3E50"
    BG_INPUT       = "#F4F6F8"
    PRIMARY        = "#3498DB"
    PRIMARY_HOVER  = "#2980B9"
    SUCCESS        = "#2ECC71"
    SUCCESS_DARK   = "#27AE60"
    DANGER         = "#E74C3C"
    INFO           = "#1ABC9C"
    INFO_DARK      = "#16A085"
    TEXT_PRIMARY   = "#2C3E50"
    TEXT_SECONDARY = "#5D6D7E"
    TEXT_MUTED     = "#95A5A6"
    BORDER         = "#D5D8DC"
    DIVIDER        = "#E8EAED"


C = Colors if _T else _C


# ── Style Treeview ────────────────────────────────────────────────────────────
def _apply_tree_style():
    s = ttk.Style()
    try:
        s.theme_use("clam")
    except Exception:
        pass
    s.configure("StockFrs.Treeview",
                 background=C.BG_CARD, foreground=C.TEXT_PRIMARY,
                 fieldbackground=C.BG_CARD, rowheight=26,
                 font=("Segoe UI", 10),
                 borderwidth=0)
    s.configure("StockFrs.Treeview.Heading",
                 background=C.BG_HEADER, foreground="#FFFFFF",
                 font=("Segoe UI", 10, "bold"),
                 relief="flat", padding=(6, 4))
    s.map("StockFrs.Treeview",
          background=[("selected", C.PRIMARY)],
          foreground=[("selected", "#FFFFFF")])


# ====================================================================
# PageStockParFrs
# ====================================================================

class PageStockParFrs(ctk.CTkFrame):

    COLONNES = ("Code Article", "Désignation", "Unité", "Stock Actuel", "Dernier Fournisseur", "Prix d'achat")

    def __init__(self, master, db_conn=None, session_data=None, iduser=None):
        super().__init__(master, fg_color=C.BG_PAGE)
        if iduser is not None:
            self.iduser = iduser
        elif session_data and 'user_id' in session_data:
            self.iduser = session_data['user_id']
        else:
            self.iduser = 1

        self._frs_map = {}   # {nom_frs: idfrs}
        self.all_data = []

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        _apply_tree_style()
        self._setup_ui()
        self.after(50, self._charger_fournisseurs)

    # ── font helper ──────────────────────────────────────────────────────────
    def _f(self, size=11, weight="normal"):
        return ctk.CTkFont(family="Segoe UI", size=size, weight=weight)

    # ====================================================================
    # UI
    # ====================================================================
    def _setup_ui(self):

        # ── En-tête ───────────────────────────────────────────────────────
        hdr = ctk.CTkFrame(self, fg_color=C.BG_HEADER, corner_radius=0, height=44)
        hdr.grid(row=0, column=0, sticky="ew")
        hdr.grid_propagate(False)
        ctk.CTkLabel(
            hdr, text="📦  Stock par Fournisseur",
            font=self._f(16, "bold"), text_color="#FFFFFF"
        ).pack(side="left", padx=16, pady=10)

        # ── Barre de filtres ──────────────────────────────────────────────
        panel = ctk.CTkFrame(self, fg_color=C.BG_CARD, corner_radius=8)
        panel.grid(row=1, column=0, sticky="ew", padx=12, pady=6)

        inner = ctk.CTkFrame(panel, fg_color="transparent")
        inner.pack(fill="x", padx=10, pady=8)

        # Label fournisseur
        ctk.CTkLabel(
            inner, text="Fournisseur :",
            font=self._f(11, "bold"), text_color=C.TEXT_PRIMARY
        ).pack(side="left", padx=(0, 6))

        # ── ComboBox fournisseurs ─────────────────────────────────────────
        self.combo_frs = ctk.CTkComboBox(
            inner,
            values=["Chargement…"],
            width=320, height=32,
            fg_color=C.BG_INPUT,
            border_color=C.BORDER,
            button_color=C.PRIMARY,
            button_hover_color=C.PRIMARY_HOVER,
            text_color=C.TEXT_PRIMARY,
            font=self._f(11),
            state="readonly",
            command=self._on_frs_change
        )
        self.combo_frs.pack(side="left", padx=(0, 10))

        # Bouton Actualiser
        ctk.CTkButton(
            inner, text="🔄 Actualiser",
            command=self._actualiser,
            fg_color=C.PRIMARY, hover_color=C.PRIMARY_HOVER,
            text_color="#FFFFFF",
            height=32, width=110, font=self._f(10, "bold")
        ).pack(side="left", padx=(0, 4))

        # Indicateur de chargement
        self.label_chargement = ctk.CTkLabel(
            inner, text="",
            font=self._f(10), text_color=C.TEXT_MUTED)
        self.label_chargement.pack(side="left", padx=(8, 0))

        # Export Excel (droite)
        ctk.CTkButton(
            inner, text="📊  Export Excel",
            command=self._exporter_excel,
            fg_color=C.INFO_DARK, hover_color=C.INFO,
            text_color="#FFFFFF",
            height=32, width=145, font=self._f(10, "bold")
        ).pack(side="right")

        # ── Zone Treeview ─────────────────────────────────────────────────
        tree_card = ctk.CTkFrame(self, fg_color=C.BG_CARD, corner_radius=8)
        tree_card.grid(row=2, column=0, sticky="nsew", padx=12, pady=(0, 4))
        tree_card.grid_rowconfigure(0, weight=1)
        tree_card.grid_columnconfigure(0, weight=1)
        self._tree_card = tree_card
        self.tree = None
        self._creer_treeview()

        # ── Footer ────────────────────────────────────────────────────────
        footer = ctk.CTkFrame(self, fg_color="transparent")
        footer.grid(row=3, column=0, sticky="ew", padx=12, pady=(2, 8))

        self.label_total = ctk.CTkLabel(
            footer, text="Total articles : 0",
            font=self._f(10, "bold"), text_color=C.PRIMARY)
        self.label_total.pack(side="left")

        self.label_maj = ctk.CTkLabel(
            footer, text="",
            font=self._f(9), text_color=C.TEXT_MUTED)
        self.label_maj.pack(side="right")

    # ====================================================================
    # Treeview
    # ====================================================================
    def _creer_treeview(self):
        if self.tree:
            self.tree.destroy()

        self.tree = ttk.Treeview(
            self._tree_card,
            columns=self.COLONNES,
            show="headings",
            style="StockFrs.Treeview",
            selectmode="browse")

        self.tree.tag_configure("even",            background=C.BG_CARD, foreground=C.TEXT_PRIMARY)
        self.tree.tag_configure("odd",             background="#F0F4F8", foreground=C.TEXT_PRIMARY)
        self.tree.tag_configure("zero_even",       background=C.BG_CARD, foreground="#E74C3C")
        self.tree.tag_configure("zero_odd",        background="#F0F4F8", foreground="#E74C3C")

        vsb = ctk.CTkScrollbar(self._tree_card, orientation="vertical",   command=self.tree.yview)
        hsb = ctk.CTkScrollbar(self._tree_card, orientation="horizontal",  command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew", padx=(6, 0), pady=(6, 0))
        vsb.grid(row=0, column=1, sticky="ns",  pady=(6, 0))
        hsb.grid(row=1, column=0, sticky="ew",  padx=(6, 0))

        col_cfg = {
            "Code Article":        dict(width=150, anchor="center", minwidth=100),
            "Désignation":         dict(width=290, anchor="w",      minwidth=200),
            "Unité":               dict(width=110, anchor="center", minwidth=70),
            "Stock Actuel":        dict(width=110, anchor="e",      minwidth=80),
            "Dernier Fournisseur": dict(width=190, anchor="w",      minwidth=120),
            "Prix d'achat":        dict(width=130, anchor="e",      minwidth=90),
        }
        for col in self.COLONNES:
            self.tree.column(col, **col_cfg.get(col, {}))

        from treeview_sort_utils import attach_tree_sort
        attach_tree_sort(self.tree, list(self.COLONNES), configure_columns=False)
    def _remplir_treeview(self):
        for item in self.tree.get_children():
            self.tree.delete(item)

        if not self.all_data:
            self.tree.insert("", "end",
                values=("", "Aucun article trouvé pour ce fournisseur", "", "", "", ""))
            self.label_total.configure(text="Total articles : 0")
            return

        for idx, row in enumerate(self.all_data):
            zebra = "even" if idx % 2 == 0 else "odd"
            # row[3] = Stock Actuel (index 3 après ajout de Unité)
            try:
                qt = float(str(row[3]).replace(".", "").replace(",", "."))
            except Exception:
                qt = 0.0
            tag = f"zero_{zebra}" if abs(qt) < 1e-9 else zebra
            self.tree.insert("", "end", values=row, tags=(tag,))

        self.label_total.configure(text=f"Total articles : {len(self.all_data)}")

    # ====================================================================
    # Connexion DB
    # ====================================================================
    def _connect_db(self):
        try:
            with open(get_config_path('config.json')) as f:
                cfg = json.load(f)['database']
            return psycopg2.connect(
                host=cfg['host'], user=cfg['user'],
                password=cfg['password'], database=cfg['database'],
                port=cfg['port']
            )
        except Exception as e:
            messagebox.showerror("Connexion DB", str(e))
            return None

    # ====================================================================
    # Chargement fournisseurs → remplit le ComboBox
    # ====================================================================
    def _charger_fournisseurs(self):
        conn = self._connect_db()
        if not conn:
            return
        try:
            cur = conn.cursor()
            cur.execute("""
                SELECT idfrs, nomfrs
                FROM tb_fournisseur
                WHERE deleted = 0
                ORDER BY nomfrs ASC
            """)
            rows = cur.fetchall()
            self._frs_map = {nom: idfrs for idfrs, nom in rows}
            noms = list(self._frs_map.keys())

            self.combo_frs.configure(
                values=noms if noms else ["Aucun fournisseur"],
                state="readonly"
            )
            if noms:
                self.combo_frs.set(noms[0])
                self._charger_stock_frs(self._frs_map[noms[0]])
        except Exception as e:
            messagebox.showerror("Erreur fournisseurs", str(e))
        finally:
            cur.close()
            conn.close()

    # ====================================================================
    # Callbacks
    # ====================================================================
    def _on_frs_change(self, nom_selec):
        idfrs = self._frs_map.get(nom_selec)
        if idfrs is not None:
            self._charger_stock_frs(idfrs)

    def _actualiser(self):
        nom = self.combo_frs.get()
        idfrs = self._frs_map.get(nom)
        if idfrs is not None:
            self._charger_stock_frs(idfrs)

    # ====================================================================
    # Chargement stock (thread)
    # ====================================================================
    def _charger_stock_frs(self, idfrs):
        self.label_chargement.configure(text="⏳ Chargement…")
        self.label_maj.configure(text="")
        threading.Thread(
            target=self._thread_charger,
            args=(idfrs,),
            daemon=True
        ).start()

    def _thread_charger(self, idfrs):
        data = self._requete_stock(idfrs)
        self.after(0, lambda: self._appliquer_data(data))

    # ====================================================================
    # Requête SQL — basée sur la logique de page_SuiviCommande
    # ====================================================================
    def _requete_stock(self, idfrs):
        conn = self._connect_db()
        if not conn:
            return []
        try:
            cur = conn.cursor()

            # ── Étape 1 : récupérer les articles commandés chez ce fournisseur ──
            # La colonne idfrs existe directement dans tb_commandedetail (confirmé CSV)
            cur.execute("""
                SELECT DISTINCT cd.idarticle
                FROM tb_commandedetail cd
                WHERE cd.idfrs = %s
            """, (idfrs,))
            idarticles = [r[0] for r in cur.fetchall()]

            # Fallback : chercher via tb_commande si tb_commandedetail.idfrs est vide
            if not idarticles:
                cur.execute("""
                    SELECT DISTINCT cd.idarticle
                    FROM tb_commandedetail cd
                    INNER JOIN tb_commande c ON cd.idcom = c.idcom
                    WHERE c.idfrs = %s
                """, (idfrs,))
                idarticles = [r[0] for r in cur.fetchall()]

            if not idarticles:
                return []

            # ── Étape 2 : calcul du stock — même logique que page_SuiviCommande ──
            query = """
                WITH unite_hierarchie AS (
                    SELECT idarticle, idunite, niveau, qtunite
                    FROM tb_unite
                    WHERE deleted = 0
                ),
                unite_coeff AS (
                    SELECT
                        idarticle,
                        idunite,
                        exp(
                            sum(ln(NULLIF(CASE WHEN qtunite > 0 THEN qtunite ELSE 1 END, 0)))
                            OVER (
                                PARTITION BY idarticle
                                ORDER BY niveau
                                ROWS BETWEEN UNBOUNDED PRECEDING AND CURRENT ROW
                            )
                        ) AS coeff_hierarchique
                    FROM unite_hierarchie
                ),
                base_unite_par_article AS (
                    SELECT DISTINCT ON (idarticle) idarticle, idunite
                    FROM tb_unite
                    WHERE deleted = 0
                    ORDER BY idarticle, qtunite ASC, idunite ASC
                ),
                rec AS (
                    SELECT idarticle, idunite, SUM(qtlivrefrs) AS quantite
                    FROM tb_livraisonfrs
                    WHERE deleted = 0 AND idarticle = ANY(%(ids)s)
                    GROUP BY idarticle, idunite
                ),
                ven AS (
                    SELECT vd.idarticle, vd.idunite, SUM(vd.qtvente) AS quantite
                    FROM tb_ventedetail vd
                    INNER JOIN tb_vente v ON vd.idvente = v.id AND v.deleted = 0
                    WHERE vd.deleted = 0 AND vd.idarticle = ANY(%(ids)s)
                    GROUP BY vd.idarticle, vd.idunite
                ),
                tin AS (
                    SELECT idarticle, idunite, SUM(qttransfert) AS quantite
                    FROM tb_transfertdetail
                    WHERE deleted = 0 AND idarticle = ANY(%(ids)s)
                    GROUP BY idarticle, idunite
                ),
                tout AS (
                    SELECT idarticle, idunite, SUM(qttransfert) AS quantite
                    FROM tb_transfertdetail
                    WHERE deleted = 0 AND idarticle = ANY(%(ids)s)
                    GROUP BY idarticle, idunite
                ),
                sor AS (
                    SELECT idarticle, idunite, SUM(qtsortie) AS quantite
                    FROM tb_sortiedetail
                    WHERE idarticle = ANY(%(ids)s)
                    GROUP BY idarticle, idunite
                ),
                inv AS (
                    SELECT bu.idarticle, bu.idunite, SUM(i.qtinventaire) AS quantite
                    FROM tb_inventaire i
                    INNER JOIN tb_unite u ON i.codearticle = u.codearticle
                    INNER JOIN base_unite_par_article bu
                        ON bu.idarticle = u.idarticle AND bu.idunite = u.idunite
                    WHERE bu.idarticle = ANY(%(ids)s)
                    GROUP BY bu.idarticle, bu.idunite
                ),
                avo AS (
                    SELECT ad.idarticle, ad.idunite, SUM(ad.qtavoir) AS quantite
                    FROM tb_avoir av
                    INNER JOIN tb_avoirdetail ad ON av.id = ad.idavoir
                    WHERE av.deleted = 0 AND ad.deleted = 0
                      AND ad.idarticle = ANY(%(ids)s)
                    GROUP BY ad.idarticle, ad.idunite
                ),
                mouvements AS (
                    SELECT idarticle, idunite, quantite, 'rec'  AS t FROM rec
                    UNION ALL SELECT idarticle, idunite, quantite, 'ven'  FROM ven
                    UNION ALL SELECT idarticle, idunite, quantite, 'tin'  FROM tin
                    UNION ALL SELECT idarticle, idunite, quantite, 'tout' FROM tout
                    UNION ALL SELECT idarticle, idunite, quantite, 'sor'  FROM sor
                    UNION ALL SELECT idarticle, idunite, quantite, 'inv'  FROM inv
                    UNION ALL SELECT idarticle, idunite, quantite, 'avo'  FROM avo
                ),
                solde_base AS (
                    SELECT
                        m.idarticle,
                        SUM(
                            CASE m.t
                                WHEN 'rec'  THEN  m.quantite * COALESCE(uc.coeff_hierarchique, 1)
                                WHEN 'tin'  THEN  m.quantite * COALESCE(uc.coeff_hierarchique, 1)
                                WHEN 'inv'  THEN  m.quantite * COALESCE(uc.coeff_hierarchique, 1)
                                WHEN 'avo'  THEN  m.quantite * COALESCE(uc.coeff_hierarchique, 1)
                                WHEN 'ven'  THEN -m.quantite * COALESCE(uc.coeff_hierarchique, 1)
                                WHEN 'sor'  THEN -m.quantite * COALESCE(uc.coeff_hierarchique, 1)
                                WHEN 'tout' THEN -m.quantite * COALESCE(uc.coeff_hierarchique, 1)
                                ELSE 0
                            END
                        ) AS solde_base
                    FROM mouvements m
                    LEFT JOIN unite_coeff uc
                        ON uc.idarticle = m.idarticle AND uc.idunite = m.idunite
                    GROUP BY m.idarticle
                ),
                dernier_prix AS (
                    -- Dernier prix d'achat chez ce fournisseur, par article
                    SELECT DISTINCT ON (cd.idarticle)
                           cd.idarticle,
                           cd.punitcmd,
                           f.nomfrs
                    FROM tb_commandedetail cd
                    LEFT JOIN tb_commande c   ON cd.idcom  = c.idcom
                    LEFT JOIN tb_fournisseur f ON COALESCE(cd.idfrs, c.idfrs) = f.idfrs
                    WHERE COALESCE(cd.idfrs, c.idfrs) = %(idfrs)s
                      AND cd.idarticle = ANY(%(ids)s)
                      AND cd.punitcmd > 0
                    ORDER BY cd.idarticle, c.datecom DESC NULLS LAST, cd.id DESC
                )
                SELECT DISTINCT ON (a.idarticle)
                    u.codearticle,
                    a.designation,
                    u.designationunite,
                    GREATEST(0, COALESCE(sb.solde_base, 0)
                        / NULLIF(COALESCE(uc.coeff_hierarchique, 1), 0))  AS stock,
                    COALESCE(dp.nomfrs, 'Inconnu')                        AS dernier_frs,
                    COALESCE(dp.punitcmd, 0)                              AS prix_achat
                FROM tb_article a
                INNER JOIN tb_unite u
                    ON a.idarticle = u.idarticle AND u.deleted = 0
                LEFT JOIN unite_coeff uc
                    ON uc.idarticle = u.idarticle AND uc.idunite = u.idunite
                LEFT JOIN solde_base sb ON sb.idarticle = a.idarticle
                LEFT JOIN dernier_prix dp ON dp.idarticle = a.idarticle
                WHERE a.idarticle = ANY(%(ids)s)
                  AND a.deleted = 0
                ORDER BY a.idarticle, u.codearticle DESC
            """
            cur.execute(query, {"idfrs": idfrs, "ids": idarticles})
            rows = cur.fetchall()

            result = []
            for code, desig, unite, stock, frs, prix in rows:
                result.append((
                    code,
                    desig,
                    unite,
                    self._fmt(max(0, float(stock or 0))),
                    frs,
                    self._fmt(prix),
                ))
            return result

        except Exception as e:
            self.after(0, lambda: messagebox.showerror("Erreur SQL", str(e)))
            import traceback; traceback.print_exc()
            return []
        finally:
            cur.close()
            conn.close()

    def _appliquer_data(self, data):
        self.all_data = data
        self._remplir_treeview()
        self.label_chargement.configure(text="")
        self.label_maj.configure(
            text=f"Actualisé : {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")

    # ====================================================================
    # Utilitaires
    # ====================================================================
    def _fmt(self, nombre):
        try:
            return f"{float(nombre):,.2f}".replace(',', ' ').replace('.', ',').replace(' ', '.')
        except Exception:
            return "0,00"

    # ====================================================================
    # Export Excel / CSV
    # ====================================================================
    def _exporter_excel(self):
        if not self.all_data:
            messagebox.showwarning("Export", "Aucune donnée à exporter.")
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from tkinter import filedialog

            nom_frs = self.combo_frs.get() or "fournisseur"
            fichier = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"stock_{nom_frs}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
            if not fichier:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Stock par Fournisseur"

            ws.merge_cells("A1:F1")
            c = ws["A1"]
            c.value = f"Stock par Fournisseur — {nom_frs}"
            c.font  = Font(bold=True, size=13, color="FFFFFF")
            c.fill  = PatternFill("solid", fgColor="2C3E50")
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 28

            hf    = PatternFill("solid", fgColor="3498DB")
            hfont = Font(bold=True, color="FFFFFF")
            for ci, cn in enumerate(self.COLONNES, 1):
                cell = ws.cell(row=2, column=ci, value=cn)
                cell.font = hfont; cell.fill = hf
                cell.alignment = Alignment(horizontal="center", vertical="center")

            thin = Side(style="thin", color="D5D8DC")
            brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

            for ri, row in enumerate(self.all_data, 3):
                bg = PatternFill("solid", fgColor="FFFFFF" if ri % 2 == 1 else "F0F4F8")
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.fill = bg; cell.border = brd
                    cell.alignment = Alignment(
                        horizontal=("center" if ci in (1, 3) else
                                    "right"  if ci in (4, 6) else "left"),
                        vertical="center")

            for i, w in enumerate([18, 36, 14, 14, 24, 16], 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

            ws.freeze_panes = "A3"
            wb.save(fichier)
            messagebox.showinfo("Export réussi", f"Fichier enregistré :\n{fichier}")

        except ImportError:
            self._exporter_csv()
        except Exception as e:
            messagebox.showerror("Erreur export", str(e))

    def _exporter_csv(self):
        import csv
        from tkinter import filedialog
        nom_frs = self.combo_frs.get() or "fournisseur"
        fichier = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialfile=f"stock_{nom_frs}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
        if not fichier:
            return
        with open(fichier, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerow(self.COLONNES)
            for row in self.all_data:
                writer.writerow(row)
        messagebox.showinfo("Succès", f"Export CSV :\n{fichier}")


# ── Test standalone ───────────────────────────────────────────────────────────
if __name__ == "__main__":
    ctk.set_appearance_mode("light")
    ctk.set_default_color_theme("blue")
    app = ctk.CTk()
    app.title("iJeery — Stock par Fournisseur")
    app.geometry("1100x750")
    if _T:
        Theme.apply(app)
    app.grid_rowconfigure(0, weight=1)
    app.grid_columnconfigure(0, weight=1)
    PageStockParFrs(app, iduser=1).grid(row=0, column=0, sticky="nsew")
    app.mainloop()

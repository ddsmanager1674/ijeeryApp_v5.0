import customtkinter as ctk
from tkinter import ttk, messagebox
import psycopg2
import json
import os
from datetime import datetime

try:
    from pages.page_LivraisonClient import PageLivraisonClient
except ImportError:
    try:
        from page_LivraisonClient import PageLivraisonClient
    except ImportError:
        PageLivraisonClient = None


class PageLivraisonEnAttente(ctk.CTkFrame):
    def __init__(self, master, db_conn=None, session_data=None, iduser=None):
        super().__init__(master)

        # Gestion de l'ID utilisateur
        if iduser is not None:
            self.iduser = iduser
        elif session_data and 'user_id' in session_data:
            self.iduser = session_data['user_id']
        else:
            self.iduser = 1

        self.magasins = []
        self._data_cache = {}   # iid Treeview → (idclient, idmag, idarticle, idunite, reste)
        self.setup_ui()
        self.charger_magasins()
        self.charger_donnees()
        # Livraison auto après affichage complet de la fenêtre
        self.after(300, lambda: self.auto_livrer_hors_b(silencieux=True))

    # ── Connexion DB ──────────────────────────────────────────────────────────
    def connect_db(self):
        """Connexion via config.json (dossier exe en prod, racine projet en dev)."""
        try:
            from pages.db_helper import connect_page_db
            return connect_page_db()
        except Exception as e:
            messagebox.showerror(
                "Erreur de connexion",
                f"Impossible de se connecter à la base.\nVérifiez config.json à côté de l'exe.\n{e}",
            )
            return None

    # ── Formatage ─────────────────────────────────────────────────────────────
    def formater_nombre(self, nombre):
        """Formate les nombres pour l'affichage (ex: 1.250,00)"""
        try:
            val = float(nombre)
            if val == int(val):
                return f"{int(val):,}".replace(',', '.')
            return f"{val:,.2f}".replace(',', '\x00').replace('.', ',').replace('\x00', '.')
        except Exception:
            return str(nombre) if nombre is not None else "0"

    # ── Interface ─────────────────────────────────────────────────────────────
    def setup_ui(self):
        """Création de l'interface utilisateur"""
        # Titre
        ctk.CTkLabel(
            self,
            text="🚚 Livraisons en Attente",
            font=ctk.CTkFont(family="Segoe UI", size=20, weight="bold")
        ).pack(pady=10)

        # ── Filtres ──────────────────────────────────────────────────────────
        frame_filtres = ctk.CTkFrame(self)
        frame_filtres.pack(fill="x", padx=20, pady=10)

        ctk.CTkLabel(
            frame_filtres, text="🔍 Recherche:",
            font=ctk.CTkFont(family="Segoe UI", size=12)
        ).pack(side="left", padx=5)

        self.entry_recherche = ctk.CTkEntry(
            frame_filtres,
            placeholder_text="Code article ou Désignation...",
            width=300
        )
        self.entry_recherche.pack(side="left", padx=5)
        self.entry_recherche.bind('<KeyRelease>', self.filtrer_donnees)

        ctk.CTkLabel(
            frame_filtres, text="🏪 Magasin:",
            font=ctk.CTkFont(family="Segoe UI", size=12)
        ).pack(side="left", padx=(20, 5))

        self.combo_magasin = ctk.CTkComboBox(
            frame_filtres,
            values=["Tous"],
            width=200,
            command=self.filtrer_donnees
        )
        self.combo_magasin.pack(side="left", padx=5)
        self.combo_magasin.set("Tous")

        # Boutons droite
        ctk.CTkButton(
            frame_filtres,
            text="📊 Export Excel",
            command=self.exporter_excel,
            fg_color="#0288d1",
            width=120
        ).pack(side="right", padx=5)

        ctk.CTkButton(
            frame_filtres,
            text="🔄 Actualiser",
            command=self.charger_donnees,
            fg_color="#2e7d32",
            width=120
        ).pack(side="right", padx=5)

        ctk.CTkButton(
            frame_filtres,
            text="⚡ Auto-Livrer (Mag A)",
            command=self.auto_livrer_hors_b,
            fg_color="#e65100",
            hover_color="#bf360c",
            width=170
        ).pack(side="right", padx=5)

        ctk.CTkButton(
            frame_filtres,
            text="📦 Livrer la sélection",
            command=self.livrer_selection,
            fg_color="#6a1b9a",
            hover_color="#4a148c",
            width=170
        ).pack(side="right", padx=5)

        ctk.CTkButton(
            frame_filtres,
            text="✅ Tout sélectionner",
            command=self.selectionner_tout,
            fg_color="#37474f",
            hover_color="#263238",
            width=150
        ).pack(side="right", padx=5)

        # ── Tableau ──────────────────────────────────────────────────────────
        frame_tableau = ctk.CTkFrame(self)
        frame_tableau.pack(fill="both", expand=True, padx=20, pady=10)

        # Colonnes sans Stock Théorique ni Stock Réel
        colonnes = (
            "N° Facture",
            "Code Article",
            "Désignation Article",
            "Unité",
            "Magasin",
            "Nom Client",
            "Reste à Livrer",
        )

        self.tree = ttk.Treeview(
            frame_tableau,
            columns=colonnes,
            show="headings",
            height=20
        )

        largeurs = {
            "N° Facture":          150,
            "Code Article":        140,
            "Désignation Article": 250,
            "Unité":               100,
            "Magasin":             160,
            "Nom Client":          200,
            "Reste à Livrer":      140,
        }

        cols_gauche = {"Désignation Article", "Nom Client"}
        for col in colonnes:
            align = "w" if col in cols_gauche else "center"
            self.tree.heading(col, text=col, anchor="w" if col in cols_gauche else "center")
            self.tree.column(col, width=largeurs.get(col, 120), anchor=align)

        # Scrollbars
        scrollbar_y = ctk.CTkScrollbar(
            frame_tableau,
            orientation="vertical",
            command=self.tree.yview
        )
        scrollbar_x = ctk.CTkScrollbar(
            frame_tableau,
            orientation="horizontal",
            command=self.tree.xview
        )
        self.tree.configure(
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set
        )

        self.tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")

        frame_tableau.grid_rowconfigure(0, weight=1)
        frame_tableau.grid_columnconfigure(0, weight=1)

        # Double-clic → ouvrir la facture dans PageLivraisonClient
        self.tree.bind("<Double-1>", self._on_double_clic_ligne)

        # ── Barre d'état ─────────────────────────────────────────────────────
        frame_info = ctk.CTkFrame(self)
        frame_info.pack(fill="x", padx=20, pady=10)

        self.label_total = ctk.CTkLabel(
            frame_info,
            text="Total lignes: 0",
            font=ctk.CTkFont(family="Segoe UI", size=12, weight="bold")
        )
        self.label_total.pack(side="left", padx=20)

        self.label_maj = ctk.CTkLabel(
            frame_info,
            text="Dernière MAJ: --",
            font=ctk.CTkFont(family="Segoe UI", size=12)
        )
        self.label_maj.pack(side="right", padx=20)

    # ── Double-clic → ouvrir PageLivraisonClient en popup ───────────────────
    def _on_double_clic_ligne(self, event):
        """Ouvre une fenêtre PageLivraisonClient avec la facture déjà chargée."""
        if PageLivraisonClient is None:
            messagebox.showerror("Erreur", "Module PageLivraisonClient introuvable.")
            return
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0])["values"]
        # values : (refvente, code_art, designation, unite, magasin, nomcli, reste)
        refvente = str(values[0])
        nomcli   = str(values[5])

        # Récupérer idclient et idmag depuis la base pour cette refvente
        conn = self.connect_db()
        if not conn:
            return
        try:
            cur = conn.cursor()
            cur.execute(
                "SELECT v.idclient, v.idmag "
                "FROM tb_vente v "
                "WHERE v.refvente = %s AND v.deleted = 0 LIMIT 1",
                (refvente,)
            )
            row = cur.fetchone()
            if not row:
                messagebox.showerror("Erreur", f"Facture {refvente} introuvable.")
                return
            idclient, idmag = row
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur DB : {e}")
            return
        finally:
            cur.close()
            conn.close()

        # Ouvrir PageLivraisonClient dans une fenêtre CTkToplevel
        win = ctk.CTkToplevel(self)
        win.title(f"Livraison — {refvente} — {nomcli}")
        win.geometry("1100x650")
        win.resizable(True, True)

        page = PageLivraisonClient(win, id_user_connecte=self.iduser)
        page.pack(fill="both", expand=True)

        # Charger la facture directement
        win.after(150, lambda: page.ouvrir_depuis_attente(refvente, nomcli, idclient, idmag))

        # Forcer l'affichage au premier plan
        def _forcer():
            win.attributes("-topmost", True)
            win.lift()
            win.focus_force()
            win.after(200, lambda: win.attributes("-topmost", False))
        win.after(100, _forcer)

        # À la fermeture : actualiser le tableau
        def _on_close():
            win.grab_release()
            win.destroy()
            self.charger_donnees()
        win.protocol("WM_DELETE_WINDOW", _on_close)

    # ── Livraison automatique pour le magasin A ───────────────────────────────
    def auto_livrer_hors_b(self, silencieux=False):
        """
        Pour tous les magasins dont la désignation est 'A' :
        insère automatiquement dans tb_livraisoncli une ligne de livraison
        complète (qtlivrecli = reste) pour toute facture VALIDEE
        qui n'a pas encore été entièrement livrée.
        """
        conn = self.connect_db()
        if not conn:
            return
        try:
            cur = conn.cursor()

            cur.execute("""
                SELECT
                    v.refvente,
                    v.idclient,
                    vd.idmag,
                    vd.idarticle,
                    vd.idunite,
                    vd.qtvente,
                    COALESCE(lv_sum.total_livre, 0) AS total_livre
                FROM tb_vente v
                INNER JOIN tb_ventedetail vd ON vd.idvente = v.id
                INNER JOIN tb_magasin m ON m.idmag = vd.idmag
                LEFT JOIN (
                    SELECT refvente, idarticle, idunite, idmag,
                           SUM(qtlivrecli) AS total_livre
                    FROM tb_livraisoncli
                    GROUP BY refvente, idarticle, idunite, idmag
                ) lv_sum ON lv_sum.refvente  = v.refvente
                        AND lv_sum.idarticle = vd.idarticle
                        AND lv_sum.idunite   = vd.idunite
                        AND lv_sum.idmag     = vd.idmag
                WHERE v.deleted = 0
                  AND v.statut = 'VALIDEE'
                  AND UPPER(TRIM(m.designationmag)) = 'A'
                  AND (vd.qtvente - COALESCE(lv_sum.total_livre, 0)) > 0
            """)
            lignes = cur.fetchall()

            if not lignes:
                if not silencieux:
                    messagebox.showinfo(
                        "Auto-Livraison",
                        "Aucune livraison en attente pour le magasin A."
                    )
                return

            # Générer une référence BL automatique unique
            from datetime import datetime as _dt
            cur.execute(
                "SELECT COUNT(*) FROM tb_livraisoncli "
                "WHERE EXTRACT(YEAR FROM dateregistre) = %s",
                (_dt.now().year,)
            )
            count = cur.fetchone()[0] + 1
            refliv = f"{_dt.now().year}-BL-AUTO-{count:05d}"
            now    = _dt.now()

            # Insérer une ligne de livraison complète pour chaque reste
            data = [
                (
                    refliv,
                    refvente,
                    idmag,
                    idarticle,
                    idunite,
                    float(qtvente - total_livre),   # qtvente  = qté livrée ce coup
                    float(qtvente - total_livre),   # qtlivrecli = idem
                    now,
                    self.iduser,
                    idclient
                )
                for refvente, idclient, idmag, idarticle, idunite, qtvente, total_livre
                in lignes
            ]

            cur.executemany("""
                INSERT INTO tb_livraisoncli
                (reflivcli, refvente, idmag, idarticle, idunite,
                 qtvente, qtlivrecli, dateregistre, iduser, idclient)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, data)
            conn.commit()

            if not silencieux:
                messagebox.showinfo(
                    "Auto-Livraison",
                    f"{len(data)} ligne(s) livrée(s) automatiquement (Magasin A)."
                )
            self.charger_donnees()

        except Exception as e:
            conn.rollback()
            if not silencieux:
                messagebox.showerror("Erreur Auto-Livraison", str(e))
            print(f"Erreur auto_livrer_hors_b: {e}")
            import traceback; traceback.print_exc()
        finally:
            cur.close()
            conn.close()

    # ── Sélectionner toutes les lignes visibles ───────────────────────────────
    def selectionner_tout(self):
        """Sélectionne toutes les lignes actuellement visibles dans le tableau."""
        items_visibles = [
            item for item in self.tree.get_children()
            if self.tree.parent(item) == ''
        ]
        self.tree.selection_set(items_visibles)

    # ── Livraison en masse de la sélection ────────────────────────────────────
    def livrer_selection(self):
        """
        Insère dans tb_livraisoncli une ligne de livraison complète
        (qtlivrecli = reste_a_livrer) pour chaque ligne sélectionnée dans le tableau.
        Les IDs (idarticle, idunite, idmag, idclient) sont lus depuis _data_cache
        peuplé lors du dernier charger_donnees — aucun risque de collision de clés.
        """
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning(
                "Aucune sélection",
                "Veuillez sélectionner au moins une ligne à livrer.\n\n"
                "💡 Astuce : cliquez sur 'Tout sélectionner' pour tout prendre."
            )
            return

        nb = len(selection)
        reponse = messagebox.askyesno(
            "Confirmer la livraison",
            f"Vous allez livrer {nb} ligne(s) sélectionnée(s).\n\n"
            "La quantité livrée sera égale au 'Reste à Livrer' pour chaque ligne.\n\n"
            "Continuer ?"
        )
        if not reponse:
            return

        conn = self.connect_db()
        if not conn:
            return

        try:
            cur = conn.cursor()

            # Générer la référence BL unique pour ce lot
            from datetime import datetime as _dt
            cur.execute(
                "SELECT COUNT(*) FROM tb_livraisoncli "
                "WHERE EXTRACT(YEAR FROM dateregistre) = %s",
                (_dt.now().year,)
            )
            count  = cur.fetchone()[0] + 1
            refliv = f"{_dt.now().year}-BL-SEL-{count:05d}"
            now    = _dt.now()

            data = []
            non_trouvees = []

            for iid in selection:
                v = self.tree.item(iid)['values']
                refvente = str(v[0])

                if iid not in self._data_cache:
                    non_trouvees.append(f"{refvente} / {v[1]} / {v[4]}")
                    continue

                idclient, idmag, idarticle, idunite, reste = self._data_cache[iid]
                if reste <= 0:
                    continue

                data.append((
                    refliv,
                    refvente,
                    idmag,
                    idarticle,
                    idunite,
                    reste,       # qtvente dans la livraison = reste à livrer
                    reste,       # qtlivrecli
                    now,
                    self.iduser,
                    idclient
                ))

            if not data:
                messagebox.showwarning(
                    "Rien à insérer",
                    "Toutes les lignes sélectionnées sont déjà livrées ou introuvables."
                )
                return

            cur.executemany("""
                INSERT INTO tb_livraisoncli
                (reflivcli, refvente, idmag, idarticle, idunite,
                 qtvente, qtlivrecli, dateregistre, iduser, idclient)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, data)
            conn.commit()

            msg = f"✅ {len(data)} ligne(s) insérée(s) avec succès.\nRéférence BL : {refliv}"
            if non_trouvees:
                msg += f"\n\n⚠️ {len(non_trouvees)} ligne(s) introuvable(s) :\n" + "\n".join(non_trouvees[:10])

            messagebox.showinfo("Livraison effectuée", msg)
            self.charger_donnees()

        except Exception as e:
            conn.rollback()
            messagebox.showerror("Erreur Livraison", f"Une erreur est survenue :\n{str(e)}")
            import traceback; traceback.print_exc()
        finally:
            cur.close()
            conn.close()

    # ── Chargement magasins ───────────────────────────────────────────────────
    def charger_magasins(self):
        """Charge la liste des magasins pour le combobox"""
        conn = self.connect_db()
        if not conn:
            return

        try:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT idmag, designationmag
                FROM tb_magasin
                WHERE deleted = 0
                ORDER BY designationmag
            """)
            self.magasins = cursor.fetchall()
            valeurs_combo = ["Tous"] + [mag[1] for mag in self.magasins]
            self.combo_magasin.configure(values=valeurs_combo)
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur chargement magasins: {e}")
        finally:
            cursor.close()
            conn.close()

    # ── Chargement données ────────────────────────────────────────────────────
    def charger_donnees(self):
        """Charge les livraisons clients en attente (sans calcul de stock)"""
        for item in self.tree.get_children():
            self.tree.delete(item)

        conn = self.connect_db()
        if not conn:
            return

        try:
            cursor = conn.cursor()

            query = """
                SELECT
                    v.refvente,
                    u.codearticle,
                    a.designation,
                    u.designationunite,
                    m.designationmag,
                    c.nomcli,
                    (vd.qtvente - COALESCE(lv_sum.total_livre, 0)) AS reste_a_livrer,
                    v.idclient,
                    vd.idmag,
                    vd.idarticle,
                    vd.idunite
                FROM tb_vente v
                INNER JOIN tb_ventedetail vd ON vd.idvente = v.id
                INNER JOIN tb_article  a  ON vd.idarticle = a.idarticle
                INNER JOIN tb_unite    u  ON (vd.idarticle = u.idarticle AND vd.idunite = u.idunite)
                INNER JOIN tb_magasin  m  ON vd.idmag = m.idmag
                INNER JOIN tb_client   c  ON v.idclient = c.idclient
                LEFT JOIN (
                    SELECT refvente, idarticle, idunite, idmag,
                           SUM(qtlivrecli) AS total_livre
                    FROM tb_livraisoncli
                    GROUP BY refvente, idarticle, idunite, idmag
                ) lv_sum ON lv_sum.refvente  = v.refvente
                         AND lv_sum.idarticle = vd.idarticle
                         AND lv_sum.idunite   = vd.idunite
                         AND lv_sum.idmag     = vd.idmag
                WHERE a.deleted = 0
                  AND v.deleted = 0
                  AND v.statut = 'VALIDEE'
                  AND (vd.qtvente - COALESCE(lv_sum.total_livre, 0)) > 0
                ORDER BY v.refvente DESC, u.codearticle, m.designationmag
            """

            cursor.execute(query)
            resultats = cursor.fetchall()



            self._data_cache.clear()

            for row in resultats:
                (refvente, code_art, designation, unite, magasin,
                 nom_client, reste_a_livrer,
                 idclient, idmag, idarticle, idunite) = row
                iid = self.tree.insert("", "end", values=(
                    refvente,
                    code_art,
                    designation,
                    unite,
                    magasin,
                    nom_client,
                    self.formater_nombre(reste_a_livrer),
                ))
                self._data_cache[iid] = (
                    idclient, idmag, idarticle, idunite, float(reste_a_livrer)
                )

            self.label_total.configure(text=f"Total lignes: {len(resultats)}")
            self.label_maj.configure(
                text=f"Dernière MAJ: {datetime.now().strftime('%H:%M:%S')}"
            )

        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur chargement données: {e}")
            print(f"Détails erreur: {e}")
        finally:
            cursor.close()
            conn.close()

    # ── Filtrage ──────────────────────────────────────────────────────────────
    def filtrer_donnees(self, event=None):
        """Filtre les données selon la recherche et le magasin sélectionné"""
        terme_recherche = self.entry_recherche.get().lower()
        magasin_filtre  = self.combo_magasin.get()

        # Parcourir TOUS les items (attachés + détachés) via le cache
        for iid in list(self._data_cache.keys()):
            try:
                values = self.tree.item(iid)['values']
            except Exception:
                continue

            code        = str(values[1]).lower()
            designation = str(values[2]).lower()
            magasin     = str(values[4])

            correspond_recherche = (
                terme_recherche in code or terme_recherche in designation
            ) if terme_recherche else True

            correspond_magasin = (
                magasin_filtre == "Tous" or magasin == magasin_filtre
            )

            if correspond_recherche and correspond_magasin:
                self.tree.reattach(iid, '', 'end')
            else:
                self.tree.detach(iid)

    # ── Export Excel ──────────────────────────────────────────────────────────
    def exporter_excel(self):
        """Exporte les données visibles vers un fichier Excel"""
        try:
            from tkinter import filedialog
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            items_visibles = [
                item for item in self.tree.get_children()
                if self.tree.parent(item) == ''
            ]

            if not items_visibles:
                messagebox.showwarning("Attention", "Aucune donnée à exporter")
                return

            fichier = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"livraisons_attente_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            if not fichier:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Livraisons en Attente"

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),  right=Side(style='thin'),
                top=Side(style='thin'),   bottom=Side(style='thin')
            )

            colonnes_excel = [
                "N° Facture",
                "Code Article",
                "Désignation Article",
                "Unité",
                "Magasin",
                "Nom Client",
                "Reste à Livrer",
            ]

            for col_num, colonne in enumerate(colonnes_excel, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value     = colonne
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = border

            for row_num, item in enumerate(items_visibles, 2):
                values = self.tree.item(item)['values']
                for col_num, value in enumerate(values, 1):
                    cell           = ws.cell(row=row_num, column=col_num)
                    cell.value     = value
                    cell.border    = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            largeurs_excel = {1: 20, 2: 15, 3: 35, 4: 12, 5: 20, 6: 25, 7: 18}
            for col_num, largeur in largeurs_excel.items():
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = largeur

            derniere_ligne = len(items_visibles) + 3
            ws.cell(row=derniere_ligne, column=1).value = (
                f"Exporté le: {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')}"
            )
            ws.cell(row=derniere_ligne, column=1).font = Font(italic=True, size=9)
            ws.cell(row=derniere_ligne + 1, column=1).value = f"Total lignes: {len(items_visibles)}"
            ws.cell(row=derniere_ligne + 1, column=1).font = Font(bold=True, size=10)

            wb.save(fichier)
            messagebox.showinfo(
                "Succès",
                f"Export réussi !\n\n{len(items_visibles)} lignes exportées vers:\n{fichier}"
            )

        except ImportError:
            messagebox.showerror(
                "Erreur",
                "Le module 'openpyxl' n'est pas installé.\n\nInstallez-le avec:\npip install openpyxl"
            )
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'export Excel:\n{str(e)}")


# ── Test ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = ctk.CTk()
    app.geometry("1000x650")
    app.title("Livraisons en Attente")

    page = PageLivraisonEnAttente(app, iduser=1)
    page.pack(fill="both", expand=True)

    app.mainloop()

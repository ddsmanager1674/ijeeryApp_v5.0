import customtkinter as ctk
from tkinter import ttk, messagebox
import psycopg2
import json
import os
from datetime import datetime

from app_theme import Colors, Fonts, Layout, styled
from log_utils import resolve_connected_user_id

class PageStockLivraison(ctk.CTkFrame):
    def __init__(self, master, db_conn=None, session_data=None, iduser=None):
        super().__init__(master, fg_color=Colors.BG_PAGE, corner_radius=0)
        
        self.iduser = (
            iduser
            if iduser is not None
            else resolve_connected_user_id(master=master, session_data=session_data)
        )
        
        self.magasins = []
        self.setup_ui()
        self.charger_magasins()
        self.charger_donnees()
    
    def connect_db(self):
        """Connexion via db_helper (config centralisée)."""
        try:
            from pages.db_helper import connect_page_db
            return connect_page_db()
        except Exception as e:
            messagebox.showerror("Erreur de connexion", f"Impossible de se connecter à la base.\nErreur: {e}")
            return None
    
    def formater_nombre(self, nombre):
        """Formate les nombres pour l'affichage (ex: 1.250,00)"""
        try:
            return f"{float(nombre):,.2f}".replace(',', ' ').replace('.', ',').replace(' ', '.')
        except:
            return "0,00"
    
    def setup_ui(self):
        """Création de l'interface utilisateur"""
        self.grid_rowconfigure(1, weight=0)
        self.grid_columnconfigure(0, weight=1)

        header = styled.frame(self, color="transparent")
        header.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 8))
        header.grid_columnconfigure(1, weight=1)

        title_box = styled.frame(header, color="transparent")
        title_box.grid(row=0, column=0, sticky="w")
        styled.label_heading(title_box, text="Stock Livraison", size=18).pack(anchor="w")
        styled.label_muted(
            title_box,
            text="Stock disponible et reste à livrer par article, magasin et client",
            size=11,
        ).pack(anchor="w", pady=(2, 0))

        actions = styled.frame(header, color="transparent")
        actions.grid(row=0, column=2, sticky="e")
        styled.button_secondary(
            actions,
            text="Actualiser",
            width=120,
            height=Layout.BTN_H,
            command=self.charger_donnees,
        ).pack(side="left", padx=(0, 8))
        styled.button_premium(
            actions,
            text="Export Excel",
            width=130,
            height=Layout.BTN_H,
            command=self.exporter_excel,
        ).pack(side="left")

        filter_card = styled.card(self)
        filter_card.grid(row=1, column=0, sticky="ew", padx=16, pady=(0, 10))
        filter_card.grid_columnconfigure(1, weight=1)

        styled.label_muted(filter_card, text="Recherche", size=11).grid(
            row=0, column=0, sticky="w", padx=(16, 8), pady=(14, 4)
        )
        self.entry_recherche = styled.entry(
            filter_card,
            placeholder="Code article ou désignation",
            height=Layout.INPUT_H,
        )
        self.entry_recherche.grid(row=1, column=0, sticky="ew", padx=(16, 8), pady=(0, 14))
        self.entry_recherche.bind('<KeyRelease>', self.filtrer_donnees)

        styled.label_muted(filter_card, text="Magasin", size=11).grid(
            row=0, column=1, sticky="w", padx=8, pady=(14, 4)
        )
        self.combo_magasin = styled.combobox(
            filter_card,
            values=["Tous"],
            height=Layout.INPUT_H,
            command=self.filtrer_donnees
        )
        self.combo_magasin.grid(row=1, column=1, sticky="ew", padx=8, pady=(0, 14))
        self.combo_magasin.set("Tous")

        table_card = styled.card(self)
        table_card.grid(row=2, column=0, sticky="nsew", padx=16, pady=(0, 10))
        table_card.grid_rowconfigure(0, weight=1)
        table_card.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)
        
        # Création du Treeview
        colonnes = (
            "Code Article",
            "Désignation Article",
            "Unité",
            "Magasin",
            "Nom Client",
            "Reste à Livrer",
            "Stock Théorique",
            "Stock Réel"
        )
        
        self.tree = ttk.Treeview(
            table_card,
            columns=colonnes,
            show="headings",
            height=20
        )
        self._configure_tree_style()
        
        # Configuration des colonnes
        largeurs = {
            "Code Article": 120,
            "Désignation Article": 250,
            "Unité": 100,
            "Magasin": 150,
            "Nom Client": 200,
            "Reste à Livrer": 120,
            "Stock Théorique": 130,
            "Stock Réel": 120
        }
        
        for col in colonnes:
            self.tree.column(col, width=largeurs.get(col, 120), anchor='center')
        
        from treeview_sort_utils import attach_tree_sort
        attach_tree_sort(self.tree, list(colonnes), configure_columns=False)
        # Scrollbars
        scrollbar_y = ctk.CTkScrollbar(
            table_card,
            orientation="vertical",
            command=self.tree.yview
        )
        scrollbar_x = ctk.CTkScrollbar(
            table_card,
            orientation="horizontal",
            command=self.tree.xview
        )
        
        self.tree.configure(
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set
        )
        
        # Placement du tableau et scrollbars
        self.tree.grid(row=0, column=0, sticky="nsew")
        scrollbar_y.grid(row=0, column=1, sticky="ns")
        scrollbar_x.grid(row=1, column=0, sticky="ew")
        
        info_bar = styled.card(self)
        info_bar.grid(row=3, column=0, sticky="ew", padx=16, pady=(0, 16))

        self.label_total = styled.badge(
            info_bar,
            text="Total lignes: 0",
            variant="info",
        )
        self.label_total.pack(side="left", padx=16, pady=12)

        self.label_maj = styled.label_muted(
            info_bar,
            text="Dernière MAJ: --",
            size=11,
        )
        self.label_maj.pack(side="right", padx=16, pady=12)

    def _configure_tree_style(self):
        style = ttk.Style()
        family = Fonts._family if getattr(Fonts, "_loaded", False) else "Segoe UI"
        style.configure(
            "StockLivraison.Treeview",
            background=Colors.BG_CARD,
            foreground=Colors.TEXT_PRIMARY,
            fieldbackground=Colors.BG_CARD,
            rowheight=Layout.ROW_H,
            borderwidth=0,
            font=(family, 9),
        )
        style.configure(
            "StockLivraison.Treeview.Heading",
            background=Colors.CLOUDS,
            foreground=Colors.TEXT_PRIMARY,
            font=(family, 9, "bold"),
        )
        style.map(
            "StockLivraison.Treeview",
            background=[("selected", Colors.PRIMARY)],
            foreground=[("selected", Colors.TEXT_ON_DARK)],
        )
        self.tree.configure(style="StockLivraison.Treeview")
    
    def charger_magasins(self):
        """Charge la liste des magasins pour le combobox"""
        conn = self.connect_db()
        if not conn:
            return
        
        try:
            cursor = conn.cursor()
            cursor.execute("""
                SELECT idmag, designationmag 
                FROM tb_magasin 
                WHERE deleted = 0 
                ORDER BY designationmag
            """)
            self.magasins = cursor.fetchall()
            
            # Mise à jour du combobox
            valeurs_combo = ["Tous"] + [mag[1] for mag in self.magasins]
            self.combo_magasin.configure(values=valeurs_combo)
            
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur chargement magasins: {e}")
        finally:
            cursor.close()
            conn.close()
    
    def charger_donnees(self):
        """Charge les données du tableau"""
        for item in self.tree.get_children():
            self.tree.delete(item)

        conn = self.connect_db()
        if not conn:
            return

        try:
            cursor = conn.cursor()

            # ── Étape 1 : récupérer tous les restes à livrer (même logique que page_livraisonEnAttente)
            cursor.execute("""
                SELECT
                    u.codearticle,
                    a.designation,
                    u.designationunite,
                    m.designationmag,
                    c.nomcli,
                    vd.idarticle,
                    vd.idunite,
                    vd.idmag,
                    (vd.qtvente - COALESCE(lv_sum.total_livre, 0)) AS reste_a_livrer
                FROM tb_vente v
                INNER JOIN tb_ventedetail vd ON vd.idvente = v.id
                INNER JOIN tb_article  a  ON vd.idarticle = a.idarticle
                INNER JOIN tb_unite    u  ON (vd.idarticle = u.idarticle AND vd.idunite = u.idunite)
                INNER JOIN tb_magasin  m  ON vd.idmag = m.idmag
                INNER JOIN tb_client   c  ON v.idclient = c.idclient
                LEFT JOIN (
                    SELECT refvente, idarticle, idunite, idmag,
                           SUM(qtlivrecli) AS total_livre
                    FROM tb_livraisoncli
                    GROUP BY refvente, idarticle, idunite, idmag
                ) lv_sum ON lv_sum.refvente  = v.refvente
                         AND lv_sum.idarticle = vd.idarticle
                         AND lv_sum.idunite   = vd.idunite
                         AND lv_sum.idmag     = vd.idmag
                WHERE a.deleted = 0
                  AND v.deleted = 0
                  AND v.statut  IN ('VALIDEE', 'EN_ATTENTE')
                  AND (vd.qtvente - COALESCE(lv_sum.total_livre, 0)) > 0
                ORDER BY u.codearticle, m.designationmag, c.nomcli
            """)
            resultats = cursor.fetchall()

            # ── Étape 2 : calculer le stock théorique pour chaque (idarticle, idunite, idmag) unique
            # On collecte les combinaisons uniques pour éviter de recalculer plusieurs fois
            combos_uniques = list({(r[5], r[6], r[7]) for r in resultats})

            stocks = {}
            for idarticle, idunite, idmag in combos_uniques:
                stocks[(idarticle, idunite, idmag)] = self._calculer_stock_sql(
                    cursor, idarticle, idunite, idmag
                )

            # ── Étape 3 : remplir le tableau
            for row in resultats:
                code_art, designation, unite, magasin, nom_client = row[0], row[1], row[2], row[3], row[4]
                idarticle, idunite, idmag = row[5], row[6], row[7]
                reste = float(row[8])

                stock_theo = stocks.get((idarticle, idunite, idmag), 0.0)
                stock_reel = stock_theo + reste   # stock avant livraison

                self.tree.insert("", "end", values=(
                    code_art,
                    designation,
                    unite,
                    magasin,
                    nom_client,
                    self.formater_nombre(reste),
                    self.formater_nombre(stock_theo),
                    self.formater_nombre(stock_reel),
                ))

            self.label_total.configure(text=f"Total lignes: {len(resultats)}")
            self.label_maj.configure(text=f"Dernière MAJ: {datetime.now().strftime('%H:%M:%S')}")

        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur chargement données: {e}")
            print(f"Détails erreur: {e}")
            import traceback; traceback.print_exc()
        finally:
            cursor.close()
            conn.close()

    def _calculer_stock_sql(self, cursor, idarticle, idunite, idmag):
        """Stock théorique aligné sur StockManager (snapshot centralisé)."""
        try:
            from stock_service import get_snapshot_cached, stock_unite
            conn = getattr(cursor, "connection", None)
            snap = get_snapshot_cached(int(idmag), conn=conn)
            return float(stock_unite(snap, int(idarticle), int(idunite)))
        except Exception as e:
            print(f"Erreur calcul stock: {e}")
            return 0.0

    def filtrer_donnees(self, event=None):
        """Filtre les données selon la recherche et le magasin sélectionné"""
        terme_recherche = self.entry_recherche.get().lower()
        magasin_filtre = self.combo_magasin.get()
        
        for item in self.tree.get_children():
            values = self.tree.item(item)['values']
            
            # Filtrage par recherche
            code = str(values[0]).lower()
            designation = str(values[1]).lower()
            correspond_recherche = (terme_recherche in code or terme_recherche in designation) if terme_recherche else True
            
            # Filtrage par magasin
            magasin = str(values[3])
            correspond_magasin = (magasin_filtre == "Tous" or magasin == magasin_filtre)
            
            # Afficher/masquer l'item
            if correspond_recherche and correspond_magasin:
                self.tree.reattach(item, '', 'end')
            else:
                self.tree.detach(item)
    
    def exporter_excel(self):
        """Exporte les données visibles vers un fichier Excel"""
        try:
            from tkinter import filedialog
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Vérifier s'il y a des données
            items_visibles = [item for item in self.tree.get_children() if self.tree.parent(item) == '']
            
            if not items_visibles:
                messagebox.showwarning("Attention", "Aucune donnée à exporter")
                return
            
            # Demander où enregistrer
            fichier = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"stock_livraisons_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not fichier:
                return
            
            # Créer le classeur Excel
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Stock et Livraisons"
            
            # Styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # En-têtes
            colonnes = [
                "Code Article",
                "Désignation Article",
                "Unité",
                "Magasin",
                "Nom Client",
                "Reste à Livrer",
                "Stock Théorique",
                "Stock Réel"
            ]
            
            for col_num, colonne in enumerate(colonnes, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = colonne
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = border
            
            # Données
            for row_num, item in enumerate(items_visibles, 2):
                values = self.tree.item(item)['values']
                
                for col_num, value in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ajuster la largeur des colonnes
            largeurs = {
                1: 15,  # Code Article
                2: 35,  # Désignation
                3: 12,  # Unité
                4: 20,  # Magasin
                5: 25,  # Nom Client
                6: 18,  # Reste à Livrer
                7: 18,  # Stock Théorique
                8: 15   # Stock Réel
            }
            
            for col_num, largeur in largeurs.items():
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = largeur
            
            # Ajouter des informations supplémentaires
            derniere_ligne = len(items_visibles) + 3
            ws.cell(row=derniere_ligne, column=1).value = f"Exporté le: {datetime.now().strftime('%d/%m/%Y à %H:%M:%S')}"
            ws.cell(row=derniere_ligne, column=1).font = Font(italic=True, size=9)
            
            ws.cell(row=derniere_ligne + 1, column=1).value = f"Total lignes: {len(items_visibles)}"
            ws.cell(row=derniere_ligne + 1, column=1).font = Font(bold=True, size=10)
            
            # Enregistrer
            wb.save(fichier)
            
            messagebox.showinfo(
                "Succès",
                f"Export réussi !\n\n{len(items_visibles)} lignes exportées vers:\n{fichier}"
            )
            
        except ImportError:
            messagebox.showerror(
                "Erreur",
                "Le module 'openpyxl' n'est pas installé.\n\nInstallez-le avec:\npip install openpyxl"
            )
        except Exception as e:
            messagebox.showerror("Erreur", f"Erreur lors de l'export Excel:\n{str(e)}")


# Test de la fenêtre
if __name__ == "__main__":
    app = ctk.CTk()
    app.geometry("1200x700")
    app.title("Gestion Stock et Livraisons")
    
    page = PageStockLivraison(app, iduser=1)
    page.pack(fill="both", expand=True)
    
    app.mainloop()

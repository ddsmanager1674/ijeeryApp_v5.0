# -*- coding: utf-8 -*-
import os
import time
from datetime import datetime
from tkinter import filedialog, messagebox
import tkinter as tk

import pandas as pd
import customtkinter as ctk
from tkinter import ttk

from app_theme import styled, Fonts, Colors

def _try_import_openpyxl():
    """Import openpyxl à la demande pour éviter de casser l'import global."""
    try:
        from openpyxl import Workbook  # type: ignore
        from openpyxl.styles import Font, Alignment, PatternFill  # type: ignore
        return Workbook, Font, Alignment, PatternFill
    except Exception:
        return None, None, None, None


def create_export_excel_button(parent, command, text="Excel",
                               width=170, height=36, icon="📊", **kwargs):
    """
    Bouton Excel standardise avec une icone visible.
    """
    return styled.button_premium(
        parent,
        text=text,
        icon=icon,
        width=width,
        height=height,
        font=Fonts.button(14),
        **kwargs
    )


def export_dataframe_to_excel(df, filename_prefix="Export", sheet_name="Donnees",
                              parent=None, initial_dir=None):
    """
    Export centralise d'un DataFrame vers Excel avec boite de dialogue.
    """
    if df is None or len(df) == 0:
        messagebox.showwarning("Export Excel", "Aucune donnee a exporter.", parent=parent)
        return None

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    initialfile = f"{filename_prefix}_{timestamp}.xlsx"
    if initial_dir is None:
        initial_dir = os.path.expanduser("~")

    file_path = filedialog.asksaveasfilename(
        parent=parent,
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialdir=initial_dir,
        initialfile=initialfile,
        title="Enregistrer l'export Excel"
    )
    if not file_path:
        return None

    try:
        df.to_excel(file_path, index=False, sheet_name=sheet_name)
        messagebox.showinfo("Export Excel", f"Export reussi :\n{file_path}", parent=parent)
        return file_path
    except Exception as exc:
        messagebox.showerror("Erreur Export Excel", str(exc), parent=parent)
        return None


def export_treeview_to_excel(treeview, filename_prefix="Export",
                             sheet_name="Donnees", parent=None):
    """
    Export generique d'un ttk.Treeview (donnees visibles) vers Excel.
    """
    try:
        items = treeview.get_children()
    except Exception:
        items = []

    if not items:
        messagebox.showwarning("Export Excel", "Aucune donnee a exporter.", parent=parent)
        return None

    cols = list(treeview["columns"])
    headers = []
    for col in cols:
        try:
            headers.append(treeview.heading(col, "text") or col)
        except Exception:
            headers.append(col)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    initialfile = f"{filename_prefix}_{timestamp}.xlsx"
    file_path = filedialog.asksaveasfilename(
        parent=parent,
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        initialfile=initialfile,
        title="Enregistrer l'export Excel"
    )
    if not file_path:
        return None

    try:
        Workbook, Font, Alignment, PatternFill = _try_import_openpyxl()
        if Workbook is None:
            messagebox.showerror(
                "Export Excel",
                "L'export Excel nécessite le module 'openpyxl' (non installé).\n"
                "Installez-le puis relancez l'application.",
                parent=parent,
            )
            return None
        try:
            from openpyxl.utils import get_column_letter  # type: ignore
        except Exception:
            get_column_letter = None

        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        ws.append(headers)

        header_fill = PatternFill(start_color="2C3E50",
                                  end_color="2C3E50", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for item in items:
            values = treeview.item(item, "values")
            if values:
                ws.append(list(values))

        # Ajuste largeurs selon contenu
        for i, col in enumerate(ws.columns, 1):
            max_len = 0
            for cell in col:
                try:
                    v = "" if cell.value is None else str(cell.value)
                    if len(v) > max_len:
                        max_len = len(v)
                except Exception:
                    pass
            col_letter = get_column_letter(i) if get_column_letter else chr(64 + i)
            ws.column_dimensions[col_letter].width = min(60, max(10, max_len + 2))

        wb.save(file_path)
        messagebox.showinfo("Export Excel", f"Export reussi :\n{file_path}", parent=parent)
        return file_path
    except Exception as exc:
        messagebox.showerror("Erreur Export Excel", str(exc), parent=parent)
        return None


class HoverExportExcelIcon:
    """
    Bouton d'export Excel ancre sur l'entete d'un Treeview.
    """
    def __init__(self, container, command, *, colors=None,
                 text="⬇ Excel", size=15, tooltip="Excel",
                 label_text=None, treeview=None, heading_column=None,
                 badge=True, always_visible=True):
        self.container = container
        self.command = command
        self.colors = colors or Colors
        self.tooltip = tooltip
        self.treeview = treeview
        self.heading_column = heading_column
        self.always_visible = always_visible
        self._context_menu = None

        self._hide_job = None
        self._visible = False
        self._lock_visible = False
        self._last_heading_ts = 0.0

        self._frame = ctk.CTkFrame(
            container,
            width=40 if badge else size + 8 + 44,
            height=20 if badge else size + 8,
            corner_radius=13 if badge else 16,
            fg_color="transparent" if badge else self.colors.BG_CARD,
            border_width=0 if badge else 1,
            border_color=self.colors.BORDER
        )
        self._button = ctk.CTkButton(
            self._frame,
            text=text,
            width=40 if badge else size,
            height=20 if badge else size,
            command=self.command,
            fg_color=self.colors.SUCCESS_DARK,
            hover_color=self.colors.SUCCESS,
            text_color="#FFFFFF",
            corner_radius=10 if badge else 16,
            font=ctk.CTkFont(family="Roboto", size=9 if badge else 14, weight="bold")
        )
        self._button.pack(side="left", padx=(6, 4), pady=4)
        self._button.bind("<ButtonPress-1>", self._lock_show, add="+")
        self._button.bind("<ButtonRelease-1>", self._unlock_show, add="+")
        self._label = None
        if label_text:
            self._label = ctk.CTkLabel(
                self._frame,
                text=label_text,
                font=ctk.CTkFont(family="Roboto", size=11, weight="bold"),
                text_color=self.colors.TEXT_PRIMARY
            )
            self._label.pack(side="left", padx=(0, 8))
        self._frame.place_forget()

        bind_targets = [self.container, self._frame, self._button]
        if self._label is not None:
            bind_targets.append(self._label)
        for w in bind_targets:
            w.bind("<Enter>", self._on_enter, add="+")
            w.bind("<Leave>", self._on_leave, add="+")

        # Survol uniquement des entetes (si treeview fourni)
        if self.treeview is not None:
            self.treeview.bind("<Motion>", self._on_tree_motion, add="+")
            self.treeview.bind("<Leave>", self._on_tree_leave, add="+")
            self.treeview.bind("<Configure>", self._on_tree_configure, add="+")
            self.treeview.bind("<Map>", self._on_tree_configure, add="+")
            self.treeview.bind("<Button-3>", self._on_tree_heading_context, add="+")
            self.treeview.bind("<Button-2>", self._on_tree_heading_context, add="+")
            self.container.after_idle(self._show)

    def bind_to(self, *widgets):
        for w in widgets:
            w.bind("<Enter>", self._on_enter, add="+")
            w.bind("<Leave>", self._on_leave, add="+")

    def _show(self):
        self._place_above_heading()
        self._frame.lift()
        if self._visible:
            return
        self._visible = True

    def _hide(self):
        if self.always_visible:
            return
        if not self._visible:
            return
        self._frame.place_forget()
        self._visible = False

    def _on_enter(self, event=None):
        if self.treeview is not None:
            # Autorise le survol du badge lui-meme
            try:
                w = event.widget if event else None
                if w not in (self._frame, self._button, getattr(self, "_label", None)):
                    return
            except Exception:
                return
        if self._hide_job:
            self.container.after_cancel(self._hide_job)
            self._hide_job = None
        self._show()

    def _on_leave(self, event=None):
        if self.always_visible:
            return
        if self.treeview is not None:
            # Ne masque pas si la souris reste sur le badge
            self._schedule_hide_from_tree()
            return
        if self._hide_job:
            self.container.after_cancel(self._hide_job)
        self._hide_job = self.container.after(120, self._hide_if_outside)

    def _on_tree_motion(self, event=None):
        if self.always_visible:
            self._show()
            return
        try:
            region = self.treeview.identify_region(event.x, event.y)
        except Exception:
            region = None
        if self._pointer_on_badge():
            if self._hide_job:
                self.container.after_cancel(self._hide_job)
                self._hide_job = None
            self._show()
            return
        if region == "heading":
            # Optionnel: n'afficher que sur l'entête de la dernière colonne
            if self.heading_column:
                try:
                    col_ids = list(self.treeview["columns"])
                    if self.heading_column in col_ids:
                        want_idx = col_ids.index(self.heading_column) + 1
                        current = self.treeview.identify_column(event.x)
                        if current != f"#{want_idx}":
                            self._schedule_hide_from_tree()
                            return
                except Exception:
                    pass
            self._last_heading_ts = time.monotonic()
            if self._hide_job:
                self.container.after_cancel(self._hide_job)
                self._hide_job = None
            self._show()
        else:
            # Laisse un petit delai pour passer de l'entete au badge
            if time.monotonic() - self._last_heading_ts < 0.5:
                return
            self._schedule_hide_from_tree()

    def _on_tree_leave(self, event=None):
        if self.always_visible:
            return
        self._schedule_hide_from_tree()

    def _on_tree_configure(self, event=None):
        if self.always_visible:
            self.container.after_idle(self._show)

    def _on_tree_heading_context(self, event=None):
        if event is None or self.treeview is None:
            return
        try:
            if self.treeview.identify_region(event.x, event.y) != "heading":
                return
        except Exception:
            return

        if self._context_menu is None:
            self._context_menu = tk.Menu(self.treeview, tearoff=0)
            self._context_menu.add_command(
                label="Exporter en excel",
                command=self.command,
            )

        try:
            self._show()
            self._context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            try:
                self._context_menu.grab_release()
            except Exception:
                pass

    def _schedule_hide_from_tree(self):
        if self._hide_job:
            self.container.after_cancel(self._hide_job)
        if self._lock_visible:
            return
        self._hide_job = self.container.after(250, self._hide_if_not_on_badge)

    def _hide_if_not_on_badge(self):
        self._hide_job = None
        if self._lock_visible:
            return
        try:
            px = self.container.winfo_pointerx()
            py = self.container.winfo_pointery()
            w = self.container.winfo_containing(px, py)
            if w in (self._frame, self._button, getattr(self, "_label", None)):
                return
        except Exception:
            pass
        self._hide()

    def _pointer_on_badge(self):
        try:
            px = self.container.winfo_pointerx()
            py = self.container.winfo_pointery()
            w = self.container.winfo_containing(px, py)
            return w in (self._frame, self._button, getattr(self, "_label", None))
        except Exception:
            return False

    def _lock_show(self, event=None):
        self._lock_visible = True
        if self._hide_job:
            self.container.after_cancel(self._hide_job)
            self._hide_job = None
        self._show()

    def _unlock_show(self, event=None):
        self._lock_visible = False

    def _hide_if_outside(self):
        self._hide_job = None
        if not self._pointer_inside_container():
            self._hide()

    def _pointer_inside_container(self):
        try:
            x = self.container.winfo_rootx()
            y = self.container.winfo_rooty()
            w = self.container.winfo_width()
            h = self.container.winfo_height()
            px = self.container.winfo_pointerx()
            py = self.container.winfo_pointery()
            return (x <= px <= x + w) and (y <= py <= y + h)
        except Exception:
            return False

    def _place_above_heading(self):
        # Placement absolu dans le Treeview, aligne sur le bord droit
        # de la derniere colonne d'entete.
        if not self.treeview:
            self._frame.place(relx=0.5, rely=0.5, anchor="center")
            return
        try:
            tree_width = int(self.treeview.winfo_width() or 0)
            if tree_width <= 1:
                # Pas encore rendu -> réessayer au prochain tick
                self.container.after(30, self._place_above_heading)
                return

            columns = self._visible_columns()
            total_width = sum(max(0, int(self.treeview.column(col, "width") or 0))
                              for col in columns)
            if total_width <= 0:
                x = tree_width - 6
            else:
                try:
                    left_fraction = float(self.treeview.xview()[0])
                except Exception:
                    left_fraction = 0.0
                scroll_offset = int(total_width * left_fraction)
                last_col_right = total_width - scroll_offset
                x = min(tree_width - 6, max(24, last_col_right - 6))

            self._frame.place(in_=self.treeview, x=x, y=2, anchor="ne")
        except Exception:
            self._frame.place(relx=0.5, rely=0.5, anchor="center")

    def _visible_columns(self):
        try:
            display_columns = self.treeview["displaycolumns"]
            all_columns = list(self.treeview["columns"])
            if not display_columns or display_columns == "#all":
                return all_columns
            columns = []
            for col in display_columns:
                if isinstance(col, int):
                    idx = col - 1
                    if 0 <= idx < len(all_columns):
                        columns.append(all_columns[idx])
                elif col in all_columns:
                    columns.append(col)
            return columns or all_columns
        except Exception:
            return list(self.treeview["columns"])


def enable_treeview_export_badge(*, colors=None, text="⬇Excel",
                                 size=15, tooltip="Excel"):
    """
    Active un bouton d'export pour tous les ttk.Treeview crees apres appel.
    """
    if getattr(ttk.Treeview, "_ijeery_export_patched", False):
        return

    colors = colors or Colors
    original_init = ttk.Treeview.__init__

    def _init(self, *args, **kwargs):
        original_init(self, *args, **kwargs)
        try:
            if hasattr(self, "_export_badge"):
                return
            cols = list(self["columns"])
            if not cols:
                return
            # Le badge est placé dans le Treeview lui-même (robuste),
            # mais on conserve un "parent" pour les after().
            container = self
            self._export_badge = HoverExportExcelIcon(
                container=container,
                command=lambda tv=self: export_treeview_to_excel(
                    tv, filename_prefix="Export", sheet_name="Donnees",
                    parent=container.winfo_toplevel() if container else None),
                colors=colors,
                text=text,
                size=size,
                tooltip=tooltip,
                label_text=None,
                treeview=self,
                heading_column=cols[-1],
                always_visible=True
            )
            self._export_badge.bind_to(self)
        except Exception:
            pass

    ttk.Treeview.__init__ = _init
    ttk.Treeview._ijeery_export_patched = True


# Activation globale
enable_treeview_export_badge()
